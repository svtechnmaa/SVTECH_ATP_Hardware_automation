import yaml
import streamlit as st
from openpyxl import load_workbook
import os
import logging
import shutil
import threading
from datetime import datetime, timedelta
import sqlite3
import pandas as pd
import json
from glob import glob
# from sacred import Experiment
# from sacred.observers import FileStorageObserver, SqlObserver
# ex = Experiment("provision_pipeline")
# sql_observer = SqlObserver('sqlite:///database.sqlite')

def read_conf():
    with open(os.environ.get("CONFIG_FILE"), "r") as f:
        config = yaml.safe_load(f)
    return config

def create_input_component(input_vars, phase, values={}):
    for name, config in input_vars.items():
        key=f"phase_{phase}_{name}"
        match config['widget']:
            case 'file_uploader':
                if values:
                    st.text_input(config['label'],value=values[name].name if hasattr(values[name], "name") else os.path.basename(values[name]), disabled=True)
                else:
                    st.session_state[f'input_data_phase_{phase}'][name] = st.file_uploader(config['label'], type=config['accept_value'], key=key)
            case 'selectbox':
                if values:
                    st.session_state[f'input_data_phase_{phase}'][name] = st.selectbox(config['label'], options=[values[name]], key=key, disabled=True)
                else:
                    if f'{phase}_{name}_options' not in st.session_state:
                        st.session_state[f'{phase}_{name}_options'] = config['accept_value']
                    index=0
                    if 'default' in config:
                        index=st.session_state[f'{phase}_{name}_options'].index(config['default'])
                    st.session_state[f'input_data_phase_{phase}'][name] = st.selectbox(config['label'], options=st.session_state[f'{phase}_{name}_options'], key=key, index=index)
            case 'text_input':
                st.session_state[f'input_data_phase_{phase}'][name] = st.text_input(config['label'], value=values[name] if values else config['default'], key=key, disabled=bool(values), type=config['type'])
            case 'multiselect':
                if values:
                    st.session_state[f'input_data_phase_{phase}'][name] = st.multiselect(config['label'], options=values[name], key=key, default=values[name], disabled=True)
                else:
                    if f'{phase}_{name}_options' not in st.session_state:
                        st.session_state[f'{phase}_{name}_options'] = config['accept_value']
                    st.session_state[f'input_data_phase_{phase}'][name] = st.multiselect(config['label'], options=st.session_state[f'{phase}_{name}_options'], key=key, default=config['default'])
            case 'checkbox':
                st.session_state[f'input_data_phase_{phase}'][name] = st.checkbox(config['label'], key=key, value=values[name] if values else config['default'], disabled=bool(values))
            case 'number_input':
                st.session_state[f'input_data_phase_{phase}'][name] = st.number_input(config['label'], key=key, value=values[name] if values else config['default'], disabled=bool(values), min_value=config['min_value'], max_value=config['max_value'], step=config['step'])
            case 'text_area':
                st.session_state[f'input_data_phase_{phase}'][name] = st.text_area(config['label'], key=key, value=values[name] if values else config['default'], disabled=bool(values))
            case 'empty':
                st.session_state[f'input_data_phase_{phase}'][f'{name}_wrapper'] = st.empty()
            case _:
                st.warning(f"Unsupported widget type: {config['widget']} (field: {name})")

def extract_tar(data, output_dir):
    import tarfile
    from io import BytesIO
    with tarfile.open(fileobj=BytesIO(data), mode='r') as tar:
        tar.extractall(output_dir)

def extract_tar_gz(data, output_dir):
    import tarfile
    from io import BytesIO
    with tarfile.open(fileobj=BytesIO(data), mode='r:gz') as tar_gz:
        tar_gz.extractall(output_dir)

def extract_zip(data, output_dir):
    import zipfile
    from io import BytesIO
    with zipfile.ZipFile(BytesIO(data), 'r') as zip_file:
        zip_file.extractall(output_dir)

def extract_rar(data, output_dir):
    import rarfile
    from io import BytesIO
    with rarfile.RarFile(BytesIO(data), 'r') as rar:
        rar.extractall(output_dir)

def CREATE_EXPORT_DIR (directory = "./" ) :
    """CREATE DIR"""
    if not os.path.exists ( directory ) :
        os.makedirs ( directory )
        logging.debug ( 'Created new directory: ' + directory )
    else :
        logging.debug ( 'Directory already existed ' + directory )
    return directory

def DELETE_DIR(directory):
    if os.path.exists ( directory ):
        shutil.rmtree(path=directory)
        logging.debug ( 'Deleted directory: ' + directory )
    else:
        logging.debug ( 'Directory not existed ' + directory )

def create_sheet_components(input_vars, phase, value={}):
    for var, var_conf in input_vars.items():
        if var_conf['widget'] == 'file_uploader' and 'xlsx' in var_conf['accept_value']:
            if not value:
                update_sheet_selection(label=f'Select {var} sheet', root_key=f'input_data_phase_{phase}', file=var, sheet=f'{var}_sheet', sheet_key=f"phase_{phase}_{var}_sheet")
            else:
                with st.session_state[f'input_data_phase_{phase}'][f'{var}_sheet_wrapper']:
                    st.session_state[f'input_data_phase_{phase}'][f'{var}_sheet']=st.selectbox(f'Select {var} sheet', [value[f'{var}_sheet']], key=f"phase_{phase}_{var}_sheet", disabled=True)

def update_sheet_selection(root_key, file, sheet, label, sheet_key, value={}):
    if value:
        with st.session_state[root_key][f'{sheet}_wrapper']:
            st.selectbox(label, [value], key=sheet_key, disabled=True)
    else:
        if file in st.session_state[root_key].keys() and st.session_state[root_key][file] and st.session_state[root_key][file].name.endswith('.xlsx'):
            try:
                workbook = load_workbook(st.session_state[root_key][file])
                sheet_names = workbook.sheetnames
                with st.session_state[root_key][f'{sheet}_wrapper']:
                    st.session_state[root_key][sheet]=st.selectbox(label, sheet_names, key=sheet_key)
            except Exception as e:
                st.error(f"Error reading Excel file: {e}")

def format_duration(td: timedelta) -> str:
    total_seconds = td.total_seconds()
    hours = int(total_seconds // 3600)
    minutes = int((total_seconds % 3600) // 60)
    seconds = total_seconds % 60
    parts = []
    if hours > 0:
        parts.append(f"{hours}h")
    if minutes > 0:
        parts.append(f"{minutes}m")
    if seconds >= 1:
        parts.append(f"{int(seconds)}s")
    elif seconds > 0:
        parts.append(f"{seconds:.2f}s")
    else:
        if not parts:
            parts.append("0s")
    return ''.join(parts)

def render_phase_page(phase, vars, statistics={}, list_runs=pd.DataFrame()):
    st.title(f"ATP phase {phase}")
    if f'input_data_phase_{phase}' not in st.session_state:
        st.session_state[f'input_data_phase_{phase}'] = {}
    st.header(f"Input {phase} data")
    create_input_component(input_vars=vars, phase=phase)
    create_sheet_components(vars, phase)
    if st.button(":star2: **RUN**", use_container_width=True, disabled=st.session_state.get('running', False) or not all(st.session_state[f'input_data_phase_{phase}'].get(name) not in [None, '', []] for name in vars.keys())):
        st.session_state.logger=StreamlitLogger()
        st.session_state.running = True
        st.session_state.running_job = phase
        st.session_state.current_running=False
        st.switch_page('pages/running.py')
    st.title(f"Statistics")
    cols = st.columns(3)
    for i in range(3):
        cols[i].markdown(
            f"""
            <div style="text-align: center; background-color: white; padding: 20px; border-radius: 10px; box-shadow: 0 1px 3px rgba(0,0,0,0.1);">
                <div style="font-size: 16px; color: #666;">{list(statistics.keys())[i]}</div>
                <div style="font-size: 28px; font-weight: bold;">{list(statistics.values())[i]}</div>
            </div>
            """, unsafe_allow_html=True)
    st.title(f"Recent Run")
    for index, row in list_runs.iterrows():
        run_id = row['run_id']
        page_link = f'<a style="display: block; padding: 5px; text-decoration: none; color: black;; font-size: 1em;" href="/running?run_id={run_id}">'
        # row_content = f'<span style="display: inline-block; width: 10%;">✅ #{row["run_id"]}</span>'+f'<span style="display: inline-block; width: 70%;">{json.dumps(row["config"])}</span>'+f'<span style="display: inline-block; width: 20%;">{row["run_time"]}</span>'
        row_content = f'<div style="display: grid; grid-template-columns: 10% 60% 20%; align-items: center; border-bottom: 1px solid #eee; padding: 5px 0; font-size: 0.8em; column-gap: auto;">'
        row_content += f'<div style="text-align: left;">✅ #{row["run_id"]}</div>' if row['result'] == 1 else f'<div style="text-align: left; color: red;">❌ #{row["run_id"]}</div>'
        config_spans = ""
        for key, value in row["config"].items():
            config_spans += f'<span style="color: black;">{key}:</span> <span style="color:#7393B3;">{value}</span>, '

        row_content += f'<div style="overflow-x: auto; text-align: left; white-space: normal; word-break: break-word;">{config_spans[:-2]}</div>'
        row_content += f'<div style="text-align: right;">{row["run_time"]}</div>'
        row_content += '</div>'
        closing_link = '</a>'
        st.markdown(f"{page_link}{row_content}{closing_link}", unsafe_allow_html=True)

def show_scrollable_log(log_content: str, height_vh=60):
    return f"""
    <div id="log-box" style="
        height: {height_vh}vh;
        overflow-y: auto;
        background-color: #f9f9f9;
        border: 1px solid #ddd;
        padding: 10px;
        font-size: 0.8rem;
        font-family: monospace;
        color: #888888;
        white-space: pre-wrap;
    ">
        {log_content}
    </div>
    """

class StreamlitLogger:
    def __init__(self):
        self.lines = []
        self.lock = threading.Lock()

    def write(self, message):
        if message.strip():
            with self.lock:
                self.lines.append(
                    f"<div style='color: #888888; font-size: 0.8rem; font-family: monospace;'>{message.strip()}</div>"
                )

    def flush(self):
        pass

    def get_html(self):
        with self.lock:
            return "<br>".join(self.lines)


class TimestampStdoutWrapper:
    def __init__(self, original_stdout, streamlit_logger):
        self.original_stdout = original_stdout
        self.streamlit_logger = streamlit_logger

    def write(self, message):
        if message.strip():
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            message_with_time = f"[{timestamp}] {message.strip()}\n"
            self.original_stdout.write(message_with_time)
            self.original_stdout.flush()
            self.streamlit_logger.write(message_with_time)

    def flush(self):
        self.original_stdout.flush()

def get_list_hd(database):
    conn = sqlite3.connect(database)
    df=pd.read_sql_query("SELECT distinct(ma_HD) from checkSN" , conn)
    return df['ma_HD'].tolist()
    # st.session_state[f'{phase}_hopdong_options'] = df['ma_HD'].tolist()

def get_list_bbbg(database, hd):
    conn = sqlite3.connect(database)
    df=pd.read_sql_query("SELECT tail FROM 'BBBG' where ma_HD=(?)" , conn, params=(hd,))
    return df['tail'].unique().tolist()
    # st.session_state[f'{phase}_list_bbbg_options'] = df['tail'].unique().tolist()

def get_list_host(database, hd):
    conn = sqlite3.connect(database)
    df=pd.read_sql_query("SELECT Hostname FROM 'BBBG' where ma_HD=(?)" , conn, params=(hd,))
    return df['Hostname'].unique().tolist()
    # st.session_state[f'{phase}_hostname_options'] = df['Hostname'].unique().tolist()

def get_list_sn(database, hd, host):
    conn = sqlite3.connect(database)
    df=pd.read_sql_query("SELECT Hostname, RealSlot, TestStatus, SN, Type FROM 'checkSN' where RealSlot IS NOT NULL and TestStatus IN ('Installed','Checked without reboot', 'Checked with reboot','Checked') and Hostname=(?) and Type in ('fpc','module','lca') and ma_HD=(?)" , conn, params=(host,hd))
    df['RealSlot'] = df['RealSlot'].apply(str)
    df['host-slot'] = df['Hostname']+' - '+df['Type']+ ' '+df['SN']+' - Slot ' + df['RealSlot']+' - '+df['TestStatus']
    return df['host-slot'].unique().tolist()
    # st.session_state[f'{phase}_hostslot_options'] = df['host-slot'].unique().tolist()

def get_statistics(database, phase):
    if os.path.isfile(database):
        conn = sqlite3.connect(database)
        df=pd.read_sql_query("SELECT start_time, stop_time, result FROM 'run' where command=(?) AND start_time IS NOT NULL AND stop_time IS NOT NULL" , conn, params=(f"run_phase{phase.replace('.', '_')}",))
        # df = df.dropna(subset=['start_time', 'stop_time'])
        df['start_time'] = pd.to_datetime(df['start_time'])
        df['stop_time'] = pd.to_datetime(df['stop_time'])
        df['duration'] = df['stop_time'] - df['start_time']
        return {'Executions': len(df), 'Success rate': str(round(((df['result'] == 1).sum() / len(df) * 100) if len(df)!=0 else 0, 2))+'%', 'Average duration': format_duration(df['duration'].mean()) if not df.empty else '0s'}
    return {'Executions': 0, 'Success rate': str(0)+'%', 'Average duration': '0s'}

def clean_config(config_json):
    if pd.isna(config_json):
        return {}
    cfg = json.loads(config_json)
    cfg.pop("seed", None)
    return cfg

def get_list_run(database, phase):
    if os.path.isfile(database):
        conn = sqlite3.connect(database)
        df=pd.read_sql_query("SELECT run_id, start_time, stop_time, result, config FROM 'run' where command=(?) AND start_time IS NOT NULL AND stop_time IS NOT NULL" , conn, params=(f"run_phase{phase.replace('.', '_')}",))
        df['config'] = df['config'].apply(clean_config)
        df['start_time'] = pd.to_datetime(df['start_time'])
        df['stop_time'] = pd.to_datetime(df['stop_time'])
        df['duration'] = df['stop_time'] - df['start_time']
        df['run_time'] = df['start_time'].dt.strftime('%-m/%-d/%y %-I:%M %p').str.lower() +" in " +df['duration'].apply(format_duration)
        df = df.sort_values(by='start_time', ascending=False)  # Sort by start_time in descending order
        return df
    return pd.DataFrame()

def get_a_run(database, id):
    conn = sqlite3.connect(database)
    df=pd.read_sql_query("SELECT * FROM 'run' where run_id=(?)" , conn, params=(id,))
    return df.iloc[0]
